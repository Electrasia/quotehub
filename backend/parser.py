"""
Local PDF parser module for QuoteHub.

Phase 1 (v0.037.0): provides read-only text + table extraction via
pdfplumber and PyMuPDF. Used by the /debug/parse endpoint to let
the operator inspect what local parsers can see BEFORE wiring them
into the main processing flow.

Phase 3+ will add format_for_llm() to convert the output into a
prompt-friendly CSV-like string for the LLM normalization step.

Both parsers are wrapped to never raise — a failing parser returns
{"available": False, "error": "..."} so the endpoint can still
return results from the other parser.
"""
import time
from pathlib import Path
from typing import Any

# Per-page text cap to keep debug responses small.
# Real quotations are <5KB of text per page; 8KB is safe and shows
# the full document in almost all cases.
MAX_TEXT_CHARS_PER_PAGE = 8000

# Per-table row cap for the same reason.
MAX_TABLE_ROWS = 200

# pdfplumber table-detection strategy.
# We prefer "lines/lines" (the default) because it gives the cleanest
# table structure — cells stay whole even when they contain multi-line
# text. We fall back to "text/text" only when lines/lines fails to
# capture meaningful data (< MIN_RICH_ROWS_FOR_CLEAN_STRATEGY rich rows),
# because text/text can split multi-line cells into separate rows.
#
# Threshold rationale: a "rich row" is a row with >= 3 non-empty cells.
# - lines/lines on a clean bordered table: every data row is rich
#   (item #, brand, model, description, qty, price, total — 7 cells).
# - lines/lines on a borderless/partial table: only the header is rich
#   (the merged header cell), data rows are empty in most columns.
# We use threshold=2 so a 1-2 item page (like the last page of a
# multi-page quote) still uses the clean strategy when it works.
# The LLM normalizer in Phase 3 will use TEXT mode (which is always
# complete) anyway, so table structure is purely diagnostic.
MIN_RICH_ROWS_FOR_CLEAN_STRATEGY = 2


def _truncate(s: str, n: int) -> str:
    if len(s) <= n:
        return s
    return s[:n] + f"... [truncated, total {len(s)} chars]"


def _count_rich_rows(tables) -> int:
    """A 'rich row' is a row with >= 3 non-None cells. This indicates
    meaningful data per row, which is a better signal than raw cell
    count (lines/lines can pack 7 meaningful cells into one row, while
    text/text can spread them across multiple rows of 1-2 cells each)."""
    return sum(1 for t in tables for row in t if sum(1 for c in row if c) >= 3)


def _extract_best_tables(page) -> tuple[list, str, int]:
    """Pick the best table-extraction strategy for this page.

    Returns (tables, strategy_name, rich_rows).

    Strategy:
      1. Try "lines/lines" first (default; cleanest structure).
      2. If it yields >= MIN_RICH_ROWS_FOR_CLEAN_STRATEGY rich rows,
         return it.
      3. Otherwise also try "text/text" (works on borderless tables).
      4. Return whichever has more rich rows.
    """
    # Try the clean default first
    try:
        clean_tables = page.extract_tables(
            table_settings={"vertical_strategy": "lines", "horizontal_strategy": "lines"}
        ) or []
    except Exception:
        clean_tables = []
    clean_score = _count_rich_rows(clean_tables)
    if clean_score >= MIN_RICH_ROWS_FOR_CLEAN_STRATEGY:
        return clean_tables, "lines/lines", clean_score

    # Default failed — try the fallback
    try:
        text_tables = page.extract_tables(
            table_settings={"vertical_strategy": "text", "horizontal_strategy": "text"}
        ) or []
    except Exception:
        text_tables = []
    text_score = _count_rich_rows(text_tables)
    if text_score > clean_score:
        return text_tables, "text/text", text_score
    return clean_tables, "lines/lines", clean_score


def parse_with_pdfplumber(pdf_path: str) -> dict[str, Any]:
    """Extract text + tables with pdfplumber. Never raises."""
    out: dict[str, Any] = {
        "available": True,
        "parser": "pdfplumber",
        "time_ms": 0,
        "pages": [],
        "total_text_chars": 0,
        "total_tables": 0,
        "total_table_rows": 0,
        "table_strategies_used": [],   # one per page, for transparency
    }
    try:
        import pdfplumber  # local import so a missing dep is caught here
    except ImportError as e:
        out["available"] = False
        out["error"] = f"pdfplumber not installed: {e}"
        return out

    t0 = time.time()
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_idx, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                text = _truncate(text, MAX_TEXT_CHARS_PER_PAGE)
                tables, strat_name, score = _extract_best_tables(page)
                page_tables = []
                for t_idx, table in enumerate(tables, start=1):
                    rows = table[:MAX_TABLE_ROWS]
                    page_tables.append({
                        "table_index": t_idx,
                        "row_count": len(table),
                        "shown_rows": len(rows),
                        "rows": rows,
                    })
                out["pages"].append({
                    "page": page_idx,
                    "text": text,
                    "text_chars": len(text),
                    "tables": page_tables,
                })
                out["total_text_chars"] += len(text)
                out["total_tables"] += len(page_tables)
                out["total_table_rows"] += sum(t["row_count"] for t in page_tables)
                out["table_strategies_used"].append({
                    "page": page_idx,
                    "strategy": strat_name,
                    "rich_rows": score,
                })
    except Exception as e:
        out["available"] = False
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        out["time_ms"] = int((time.time() - t0) * 1000)
    return out


def parse_with_pymupdf(pdf_path: str) -> dict[str, Any]:
    """Extract text + per-page metadata with PyMuPDF. Never raises."""
    out: dict[str, Any] = {
        "available": True,
        "parser": "pymupdf",
        "time_ms": 0,
        "pages": [],
        "total_text_chars": 0,
    }
    try:
        import fitz  # PyMuPDF
    except ImportError as e:
        out["available"] = False
        out["error"] = f"pymupdf not installed: {e}"
        return out

    t0 = time.time()
    try:
        doc = fitz.open(pdf_path)
        try:
            for page_idx, page in enumerate(doc, start=1):
                text = page.get_text("text") or ""
                text = _truncate(text, MAX_TEXT_CHARS_PER_PAGE)
                # Block count is a rough layout signal
                try:
                    blocks = page.get_text("blocks")
                    block_count = len(blocks)
                except Exception:
                    block_count = None
                out["pages"].append({
                    "page": page_idx,
                    "text": text,
                    "text_chars": len(text),
                    "block_count": block_count,
                })
                out["total_text_chars"] += len(text)
        finally:
            doc.close()
    except Exception as e:
        out["available"] = False
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        out["time_ms"] = int((time.time() - t0) * 1000)
    return out


def parse_pdf(pdf_path: str, ocr_enabled: bool = True, use_llm_fallback: bool = True) -> dict[str, Any]:
    """Run all available parsers and return a combined debug response.

    If both pdfplumber and PyMuPDF return little or no text from a page
    (typical for scanned/image-only PDFs) and `ocr_enabled` is True,
    triggers an OCR pass on that PDF and includes the OCR result under
    parsers["ocr"]. The OCR text is also merged into pymupdf's per-page
    text so downstream code that reads pymupdf's text gets the OCR text
    when available.

    Top-level shape:
        {
          "filename": "...",
          "file_size": int,
          "num_pages": int,
          "parsers": {
            "pdfplumber": { ... },
            "pymupdf": { ... },
            "ocr": { ... }   # only if OCR was triggered
          }
        }
    """
    p = Path(pdf_path)
    if not p.exists():
        return {"error": f"File not found: {pdf_path}"}

    file_size = p.stat().st_size

    # Use pdfplumber to get page count quickly (it errors on non-PDFs).
    num_pages = 0
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            num_pages = len(pdf.pages)
    except Exception:
        try:
            import fitz
            doc = fitz.open(pdf_path)
            num_pages = len(doc)
            doc.close()
        except Exception as e:
            return {
                "filename": p.name,
                "file_size": file_size,
                "error": f"Could not open PDF with either parser: {e}",
            }

    result = {
        "filename": p.name,
        "file_size": file_size,
        "num_pages": num_pages,
        "parsers": {
            "pdfplumber": parse_with_pdfplumber(pdf_path),
            "pymupdf": parse_with_pymupdf(pdf_path),
        },
    }

    # OCR fallback: if both parsers return little text per page (e.g.
    # scanned PDFs), trigger OCR. We measure average per-page chars.
    # Threshold: 50 chars/page average = clearly image-only PDF.
    if ocr_enabled and num_pages > 0:
        try:
            pp_total = result["parsers"]["pdfplumber"].get("total_text_chars", 0)
            pm_total = result["parsers"]["pymupdf"].get("total_text_chars", 0)
            avg_per_page = max(pp_total, pm_total) / num_pages
            if avg_per_page < 50:
                # Run OCR synchronously for pytesseract (no LLM) and
                # asynchronously for the LLM fallback. We do this in a
                # sync context here for simplicity — pytesseract is the
                # common case, the LLM fallback is rare.
                import asyncio as _asyncio
                from .ocr import ocr_pdf
                try:
                    ocr_result = _asyncio.run(
                        ocr_pdf(pdf_path, use_llm_fallback=use_llm_fallback)
                    )
                except RuntimeError:
                    # Already inside an event loop (e.g. from FastAPI).
                    # The caller should use parse_pdf_with_ocr() instead.
                    ocr_result = {
                        "source": "skipped",
                        "text": "",
                        "error": "OCR skipped: event loop already running. "
                                 "Use parse_pdf_with_ocr() from async context.",
                        "time_ms": 0,
                    }
                result["parsers"]["ocr"] = ocr_result
                # Merge OCR text into pymupdf pages so downstream code
                # that reads pymupdf text sees the OCR output.
                if ocr_result.get("text"):
                    _merge_ocr_into_pymupdf(result, ocr_result)
        except Exception as e:
            # OCR is best-effort — never let it fail the parse
            result["parsers"]["ocr"] = {
                "source": "error",
                "text": "",
                "error": f"{type(e).__name__}: {e}",
                "time_ms": 0,
            }

    return result


def _merge_ocr_into_pymupdf(result: dict, ocr_result: dict) -> None:
    """Replace per-page text with OCR text in BOTH pdfplumber and pymupdf.

    The rules-based extractor reads text from pdfplumber (and tables
    from pdfplumber too), so for scanned PDFs the OCR text must be
    visible there. We update both parser dicts so callers don't have
    to know OCR was used.
    """
    pymu = result.get("parsers", {}).get("pymupdf", {})
    pp = result.get("parsers", {}).get("pdfplumber", {})
    ocr_pages = ocr_result.get("pages") or []
    if not ocr_pages:
        # Vision LLM doesn't return per-page text — just store the full
        # text on the first page so extract_items can still find it.
        full = ocr_result.get("text", "")
        if full:
            if pymu.get("pages"):
                pymu["pages"][0]["text"] = full
                pymu["pages"][0]["text_chars"] = len(full)
            if pp.get("pages"):
                pp["pages"][0]["text"] = full
                pp["pages"][0]["text_chars"] = len(full)
                # Also bump the total_text_chars so the merge heuristic
                # downstream can see we got text
                pp["total_text_chars"] = max(
                    pp.get("total_text_chars", 0), len(full),
                )
        return
    # Map page number → OCR text
    ocr_by_page = {p.get("page"): p.get("text", "") for p in ocr_pages}
    for parser_dict in (pymu, pp):
        for page in parser_dict.get("pages", []):
            page_no = page.get("page")
            ocr_text = ocr_by_page.get(page_no, "")
            if ocr_text:
                page["text"] = ocr_text
                page["text_chars"] = len(ocr_text)
        # Update totals so any downstream code that checks the totals sees
        # the new content
        total = sum(p.get("text_chars", 0) for p in parser_dict.get("pages", []))
        parser_dict["total_text_chars"] = total


async def _parse_pdf_with_ocr_impl(pdf_path: str, ocr_enabled: bool = True,
                                     use_llm_fallback: bool = True) -> dict[str, Any]:
    """Async implementation of parse_pdf() that uses OCR. Use this from
    FastAPI handlers to avoid the RuntimeError when an event loop
    is already running.

    Same return shape as parse_pdf().
    """
    p = Path(pdf_path)
    if not p.exists():
        return {"error": f"File not found: {pdf_path}"}

    file_size = p.stat().st_size
    num_pages = 0
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            num_pages = len(pdf.pages)
    except Exception:
        try:
            import fitz
            doc = fitz.open(pdf_path)
            num_pages = len(doc)
            doc.close()
        except Exception as e:
            return {
                "filename": p.name,
                "file_size": file_size,
                "error": f"Could not open PDF with either parser: {e}",
            }

    result = {
        "filename": p.name,
        "file_size": file_size,
        "num_pages": num_pages,
        "parsers": {
            "pdfplumber": parse_with_pdfplumber(pdf_path),
            "pymupdf": parse_with_pymupdf(pdf_path),
        },
    }

    if ocr_enabled and num_pages > 0:
        try:
            pp_total = result["parsers"]["pdfplumber"].get("total_text_chars", 0)
            pm_total = result["parsers"]["pymupdf"].get("total_text_chars", 0)
            avg_per_page = max(pp_total, pm_total) / num_pages
            if avg_per_page < 50:
                from .ocr import ocr_pdf
                ocr_result = await ocr_pdf(pdf_path, use_llm_fallback=use_llm_fallback)
                result["parsers"]["ocr"] = ocr_result
                if ocr_result.get("text"):
                    _merge_ocr_into_pymupdf(result, ocr_result)
        except Exception as e:
            result["parsers"]["ocr"] = {
                "source": "error",
                "text": "",
                "error": f"{type(e).__name__}: {e}",
                "time_ms": 0,
            }

    return result


# Public alias used by main.py and other callers
parse_pdf_with_ocr = _parse_pdf_with_ocr_impl


def parse_xlsx(xlsx_path: str) -> dict[str, Any]:
    """Extract text + tables from a .xlsx file with openpyxl.

    Each sheet becomes one "page" in the result dict, and all rows
    of that sheet become a single "table" in that page. This matches
    the pdfplumber shape, so the rules-based extractor can handle
    .xlsx files without any changes.

    Cell values are coerced to strings (matching pdfplumber's behavior).
    Merged cells: only the top-left cell of a merged range gets the
    value; other cells in the range are left as None (openpyxl's
    default behavior).
    """
    out: dict[str, Any] = {
        "available": True,
        "parser": "xlsx",
        "time_ms": 0,
        "pages": [],
        "total_text_chars": 0,
        "total_tables": 0,
        "total_table_rows": 0,
    }
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        out["available"] = False
        out["error"] = f"openpyxl not installed: {e}"
        return out

    t0 = time.time()
    try:
        # data_only=True: read the cached value, not the formula
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        try:
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    # Convert to list of strings or None
                    row_list = []
                    for cell in row:
                        if cell is None:
                            row_list.append(None)
                        else:
                            row_list.append(str(cell) if cell is not None else None)
                    # Drop fully-empty trailing rows
                    if any(c is not None and str(c).strip() for c in row_list):
                        rows.append(row_list)
                if not rows:
                    continue
                # Truncate like pdfplumber
                if len(rows) > MAX_TABLE_ROWS:
                    rows = rows[:MAX_TABLE_ROWS]
                # Build the page-text view: a CSV-ish rendering
                text_lines = []
                for r in rows:
                    cells = [str(c) if c is not None else "" for c in r]
                    text_lines.append(" | ".join(cells))
                page_text = "\n".join(text_lines)
                page_text = _truncate(page_text, MAX_TEXT_CHARS_PER_PAGE)
                table_dict = {
                    "table_index": 1,
                    "row_count": len(rows),
                    "shown_rows": len(rows),
                    "rows": rows,
                }
                out["pages"].append({
                    "page": sheet_idx,
                    "text": page_text,
                    "text_chars": len(page_text),
                    "tables": [table_dict],
                    "sheet_name": sheet_name,
                })
                out["total_text_chars"] += len(page_text)
                out["total_tables"] += 1
                out["total_table_rows"] += len(rows)
        finally:
            wb.close()
    except Exception as e:
        out["available"] = False
        out["error"] = f"{type(e).__name__}: {e}"
    finally:
        out["time_ms"] = int((time.time() - t0) * 1000)
    return out


def parse_xlsx_full(xlsx_path: str) -> dict[str, Any]:
    """Top-level parse for .xlsx files. Returns a parse_result in the
    same shape as parse_pdf() so the rules-based extractor can use it.

    The output has only one parser: 'xlsx'.
    """
    p = Path(xlsx_path)
    if not p.exists():
        return {"error": f"File not found: {xlsx_path}"}

    file_size = p.stat().st_size
    result = parse_xlsx(xlsx_path)
    if not result.get("available"):
        return {
            "filename": p.name,
            "file_size": file_size,
            "num_pages": 0,
            "error": result.get("error", "openpyxl failed"),
        }
    return {
        "filename": p.name,
        "file_size": file_size,
        "num_pages": len(result["pages"]),
        "parsers": {
            "xlsx": result,
            # Aliases for compatibility with code that reads pdfplumber/pymupdf
            "pdfplumber": result,  # extract.py reads pdfplumber.pages[].tables
            "pymupdf": result,     # also read for text
        },
    }


def parse_file(file_path: str, ocr_enabled: bool = True,
               use_llm_fallback: bool = True) -> dict[str, Any]:
    """Dispatch parser by file extension.

    Supports .pdf (via parse_pdf) and .xlsx (via parse_xlsx_full).
    Returns a parse_result dict in the unified shape consumed by
    extract_items().
    """
    p = Path(file_path)
    if not p.exists():
        return {"error": f"File not found: {file_path}"}
    suffix = p.suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(file_path, ocr_enabled=ocr_enabled,
                         use_llm_fallback=use_llm_fallback)
    if suffix == ".xlsx":
        return parse_xlsx_full(file_path)
    return {
        "error": f"Unsupported file type: {suffix}. Supported: .pdf, .xlsx",
    }


async def parse_file_with_ocr(file_path: str, ocr_enabled: bool = True,
                                use_llm_fallback: bool = True) -> dict[str, Any]:
    """Async version of parse_file() — uses OCR for PDFs only.

    Same return shape as parse_file(). For .xlsx files, no OCR is
    needed (spreadsheets always have structured data).
    """
    p = Path(file_path)
    if not p.exists():
        return {"error": f"File not found: {file_path}"}
    suffix = p.suffix.lower()
    if suffix == ".pdf":
        return await parse_pdf_with_ocr(file_path, ocr_enabled=ocr_enabled,
                                         use_llm_fallback=use_llm_fallback)
    if suffix == ".xlsx":
        return parse_xlsx_full(file_path)
    return {
        "error": f"Unsupported file type: {suffix}. Supported: .pdf, .xlsx",
    }


def format_for_llm(parse_result: dict) -> str:
    """Format a parse_pdf() result into a CSV-like string the LLM can
    read in Phase 3 to extract structured data.

    Phase 2 (v0.037.0): preview only. The text is shaped for an LLM
    prompt but no LLM call is made here. Phase 3 will design the
    actual normalization prompt.

    Structure of the output (one section per page):
        === Page N ===
        [text]
        <full text from pdfplumber>
        [pymupdf]
        <full text from pymupdf, for cross-checking>
        [tables]
        <pipe-delimited rows from pdfplumber, with row index>

    All pipes inside cell values are replaced with ' / ' to keep the
    CSV unambiguous. Truncated cells are marked.
    """
    import re as _re

    def _clean(s):
        if s is None:
            return ""
        s = str(s).replace("|", " / ").replace("\r", " ").replace("\n", " ⏎ ")
        return _re.sub(r"\s+", " ", s).strip()

    lines = []
    lines.append(f"[Document]")
    lines.append(f"filename: {parse_result.get('filename', '?')}")
    lines.append(f"pages: {parse_result.get('num_pages', 0)}")
    lines.append("")

    pp = parse_result.get("parsers", {}).get("pdfplumber", {})
    pm = parse_result.get("parsers", {}).get("pymupdf", {})

    for page_idx in range(parse_result.get("num_pages", 0)):
        page_no = page_idx + 1
        lines.append(f"=== Page {page_no} ===")
        # pdfplumber text
        pp_page = next((p for p in pp.get("pages", []) if p.get("page") == page_no), None)
        if pp_page:
            lines.append(f"[pdfplumber text — {pp_page.get('text_chars', 0)} chars]")
            lines.append(pp_page.get("text", "").rstrip())
        # pymupdf text
        pm_page = next((p for p in pm.get("pages", []) if p.get("page") == page_no), None)
        if pm_page:
            lines.append("")
            lines.append(f"[pymupdf text — {pm_page.get('text_chars', 0)} chars]")
            lines.append(pm_page.get("text", "").rstrip())
        # pdfplumber tables (CSV)
        if pp_page and pp_page.get("tables"):
            lines.append("")
            lines.append(f"[pdfplumber tables — {len(pp_page['tables'])} table(s)]")
            for t in pp_page["tables"]:
                lines.append(f"  -- table {t.get('table_index')}: "
                             f"{t.get('row_count')} rows x "
                             f"{len(t['rows'][0]) if t.get('rows') else 0} cols --")
                for r_idx, row in enumerate(t.get("rows", [])):
                    cells = [_clean(c) for c in row]
                    lines.append(f"  [{r_idx}] " + " | ".join(cells))
        lines.append("")

    return "\n".join(lines)
