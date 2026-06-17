"""
backend/routes/files.py — File upload, processing, and management endpoints.

This module handles:
    - File upload
    - PDF processing (streaming)
    - Confirmation/saving
    - Skip/delete operations
    - Export/import
"""

import hashlib
import json
import logging
import os
import shutil
import zipfile
import tempfile
import asyncio
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)


# ─── Helpers ──────────────────────────────────────────────

def _count_pages(filepath: Path) -> int:
    """Count pages (PDF) or sheets (XLSX) without a full parse."""
    suffix = filepath.suffix.lower()
    if suffix == ".pdf":
        try:
            import pdfplumber
            with pdfplumber.open(str(filepath)) as pdf:
                return len(pdf.pages)
        except Exception:
            logger.warning("pdfplumber page count failed for %s", filepath, exc_info=True)
            try:
                import fitz
                doc = fitz.open(str(filepath))
                n = len(doc)
                doc.close()
                return n
            except Exception:
                logger.warning("pymupdf page count also failed for %s", filepath, exc_info=True)
                return 0
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(filepath), read_only=True, data_only=True)
            n = len(wb.sheetnames)
            wb.close()
            return n
        except Exception:
            logger.warning("openpyxl page count failed for %s", filepath, exc_info=True)
            return 0
    return 0


def _generate_page_images(filepath: Path) -> list[str]:
    """Generate PNG images for each page/sheet of a file.
    
    Returns a list of relative URL paths (e.g. /images/{stem}/page_1.png).
    """
    from ..main import IMAGES_DIR
    
    stem = filepath.stem
    img_dir = IMAGES_DIR / stem
    img_dir.mkdir(parents=True, exist_ok=True)
    
    suffix = filepath.suffix.lower()
    pages = []
    
    if suffix == ".pdf":
        try:
            import pdfplumber
            with pdfplumber.open(str(filepath)) as pdf:
                for i, page in enumerate(pdf.pages, 1):
                    img_path = img_dir / f"page_{i}.png"
                    # Render page to PNG at 150 DPI for reasonable file size
                    pil_img = page.to_image(resolution=150)
                    pil_img.save(str(img_path))
                    pages.append(f"/images/{stem}/page_{i}.png")
            return pages
        except Exception:
            logger.warning("pdfplumber image gen failed for %s", filepath, exc_info=True)
            pass
        # Fallback: try PyMuPDF
        try:
            import fitz
            doc = fitz.open(str(filepath))
            for i in range(len(doc)):
                page = doc[i]
                pix = page.get_pixmap(dpi=150)
                img_path = img_dir / f"page_{i + 1}.png"
                pix.save(str(img_path))
                pages.append(f"/images/{stem}/page_{i + 1}.png")
            doc.close()
            return pages
        except Exception:
            logger.warning("pymupdf image gen failed for %s", filepath, exc_info=True)
            pass
    
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
            from PIL import Image, ImageDraw, ImageFont
            
            wb = load_workbook(str(filepath), read_only=True, data_only=True)
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                # Read all rows from the sheet
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
                wb.close()
                
                if not rows:
                    continue
                
                # Trim trailing empty columns — find last column with data
                last_data_col = 0
                for row in rows[:50]:
                    for col_idx in range(len(row) - 1, -1, -1):
                        if row[col_idx].strip():
                            if col_idx + 1 > last_data_col:
                                last_data_col = col_idx + 1
                            break
                
                if last_data_col == 0:
                    continue  # sheet is completely empty
                
                # Render sheet as a table-based image
                font_size = 13
                padding = 8
                min_col_width = 80
                max_col_width = 200
                border_color = "#cccccc"
                header_bg = "#f0f0f0"
                line_height = 16
                
                max_cols = last_data_col
                
                # Auto-size columns based on content
                col_widths = [min_col_width] * max_cols
                try:
                    font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", font_size)
                except Exception:
                    logger.warning("Font load failed, using default", exc_info=True)
                    font = ImageFont.load_default()
                
                def _wrap_text(text, max_width):
                    """Break text into lines that fit within max_width pixels."""
                    words = str(text).split()
                    if not words:
                        return [""]
                    lines = []
                    current = words[0]
                    for word in words[1:]:
                        test = current + " " + word
                        bbox = font.getbbox(test)
                        if bbox[2] - bbox[0] <= max_width:
                            current = test
                        else:
                            lines.append(current)
                            current = word
                    lines.append(current)
                    return lines
                
                # Pre-calculate wrapped text for all cells
                cell_lines = []  # cell_lines[row_idx][col_idx] = list of lines
                for row in rows[:50]:
                    row_lines = []
                    for col_idx, cell in enumerate(row[:max_cols]):
                        w = col_widths[col_idx]
                        max_text_w = w - padding  # 4px left + ~4px right
                        lines = _wrap_text(cell, max_text_w)
                        row_lines.append(lines)
                    cell_lines.append(row_lines)
                
                # Calculate row heights based on wrapped text
                row_heights = []
                for row_idx, row_lc in enumerate(cell_lines):
                    max_lines = 1
                    for lines in row_lc:
                        if len(lines) > max_lines:
                            max_lines = len(lines)
                    row_heights.append(max(max_lines * line_height + padding, 24))
                
                img_width = sum(col_widths) + padding * 2
                img_height = sum(row_heights) + padding * 2
                
                img = Image.new("RGB", (img_width, img_height), "white")
                draw = ImageDraw.Draw(img)
                
                y = padding
                for row_idx, row_lc in enumerate(cell_lines):
                    rh = row_heights[row_idx]
                    x = padding
                    for col_idx, lines in enumerate(row_lc):
                        w = col_widths[col_idx]
                        # Draw cell background for header row
                        if row_idx == 0:
                            draw.rectangle([x, y, x + w, y + rh], fill=header_bg)
                        # Draw cell border
                        draw.rectangle([x, y, x + w, y + rh], outline=border_color)
                        # Draw wrapped text
                        for line_idx, line in enumerate(lines):
                            draw.text((x + 4, y + 4 + line_idx * line_height), line, fill="black", font=font)
                        x += w
                    y += rh
                
                img_path = img_dir / f"page_{sheet_idx}.png"
                img.save(str(img_path))
                pages.append(f"/images/{stem}/page_{sheet_idx}.png")
            return pages
        except Exception:
            logger.warning("openpyxl image gen failed for %s", filepath, exc_info=True)
            pass
    
    return pages


def _find_file_by_id(file_id: str) -> tuple[int, dict] | None:
    """Find a file entry by its stable file_id. Returns (index, entry) or None."""
    from ..main import uploaded_files
    for i, entry in enumerate(uploaded_files):
        if entry.get("file_id") == file_id:
            return i, entry
    return None


def _find_file_by_index(file_index: int) -> tuple[int, dict] | None:
    """Find a file entry by its array index (legacy). Returns (index, entry) or None."""
    from ..main import uploaded_files
    if 0 <= file_index < len(uploaded_files):
        return file_index, uploaded_files[file_index]
    return None


def _resolve_file(file_id: str | None = None, file_index: int | None = None) -> tuple[int, dict] | None:
    """Resolve a file by file_id (preferred) or file_index (fallback)."""
    if file_id:
        return _find_file_by_id(file_id)
    if file_index is not None:
        return _find_file_by_index(file_index)
    return None

from ..auth import require_role
from ..db import get_db

router = APIRouter(tags=["files"])


# ─── Models ────────────────────────────────────────────────

class ProcessRequest(BaseModel):
    model_config = {'protected_namespaces': ()}
    file_id: str | None = None
    file_index: int | None = None
    model_source: str = "auto"
    use_llm_fallback: bool = False


class ConfirmRequest(BaseModel):
    file_id: str | None = None
    file_index: int | None = None
    data: dict


class DeleteRequest(BaseModel):
    ids: list[int]


class UpdateRequest(BaseModel):
    id: int
    data: dict


class RemoveFileRequest(BaseModel):
    file_id: str


# ─── Upload ────────────────────────────────────────────────

@router.post("/upload", dependencies=[Depends(require_role("admin", "master"))])
async def upload(files: list[UploadFile] = File(...), request: Request = None):
    """Upload PDF or XLSX files for processing."""
    from ..main import uploaded_files, UPLOAD_DIR
    from ..auth import get_current_user
    from collections import Counter
    
    user = get_current_user(request)
    username = user.get("username", "unknown") if user else "unknown"
    
    results = []
    errors = []
    
    # Compute queue size once (non-saved files only)
    pending_count = sum(1 for f in uploaded_files if f.get("status") != "saved")
    MAX_QUEUED = 50
    
    for file in files:
        if not file.filename:
            continue

        # Reject if queue would exceed the limit
        if pending_count >= MAX_QUEUED:
            owner_counts = Counter(
                f.get("uploaded_by", "unknown")
                for f in uploaded_files
                if f.get("status") != "saved"
            )
            top_user, top_count = owner_counts.most_common(1)[0]
            logger.warning("Upload rejected: queue full", extra={
                'category': 'PROCESS', 'file': file.filename,
                'username': username, 'pending': pending_count
            })
            errors.append({
                "filename": file.filename,
                "error": (
                    f"Upload queue full (limit {MAX_QUEUED}). "
                    f"User '{top_user}' has {top_count} file(s) pending."
                )
            })
            continue
        
        # Validate file extension
        ext = Path(file.filename).suffix.lower()
        if ext not in (".pdf", ".xlsx"):
            logger.warning("Upload rejected: unsupported file type", extra={
                'category': 'PROCESS', 'file': file.filename, 'error': f"Unsupported type: {ext}"
            })
            errors.append({"filename": file.filename, "error": f"Unsupported file type: {ext}"})
            continue
        
        # Read file into memory for validation before writing to disk
        content = await file.read()

        # Reject empty files
        if len(content) == 0:
            logger.warning("Upload rejected: empty file", extra={
                'category': 'PROCESS', 'file': file.filename, 'error': 'Empty file (0 bytes)'
            })
            errors.append({"filename": file.filename, "error": "Empty file"})
            continue

        # Reject oversized files (configurable, default 5 MB)
        from ..utils import get_config_data
        cfg = get_config_data()
        max_bytes = cfg.get("max_upload_size_mb", 5) * 1024 * 1024
        if len(content) > max_bytes:
            limit_mb = cfg.get("max_upload_size_mb", 5)
            logger.warning("Upload rejected: file too large", extra={
                'category': 'PROCESS', 'file': file.filename,
                'size_mb': f"{len(content) / 1024 / 1024:.1f}",
                'max_mb': limit_mb,
            })
            errors.append({"filename": file.filename,
                           "error": f"File too large ({len(content) / 1024 / 1024:.1f} MB). Maximum: {limit_mb} MB"})
            continue

        # Save file to temp directory
        filepath = UPLOAD_DIR / file.filename
        with open(filepath, "wb") as f:
            f.write(content)
        
        # Add to uploaded files list
        num_pages = _count_pages(filepath)
        file_id = uuid.uuid4().hex[:12]
        entry = {
            "file_id": file_id,
            "filename": file.filename,
            "filepath": str(filepath),
            "status": "uploaded",
            "num_pages": num_pages,
            "pages": [],
            "uploaded_by": username,
        }
        uploaded_files.append(entry)
        pending_count += 1  # Track for subsequent files in this batch
        results.append({
            "file_id": file_id,
            "filename": file.filename,
            "status": "uploaded",
            "file_index": len(uploaded_files) - 1,
            "num_pages": num_pages,
            "uploaded_by": username,
        })
        
        logger.info("File uploaded", extra={
            'category': 'PROCESS',
            'file': file.filename,
            'file_id': file_id,
            'pages': num_pages
        })
    
    return {"uploaded": len(results), "files": results, "errors": errors}


@router.post("/clear", dependencies=[Depends(require_role("admin", "master"))])
async def clear_files():
    """Clear all uploaded files and clean up files on disk."""
    from ..main import uploaded_files, IMAGES_DIR
    from pathlib import Path

    for entry in uploaded_files:
        filepath = Path(entry.get("filepath", ""))
        if filepath.is_file():
            filepath.unlink()
            img_dir = IMAGES_DIR / filepath.stem
            if img_dir.exists():
                shutil.rmtree(str(img_dir), ignore_errors=True)
    uploaded_files.clear()
    logger.info("All files cleared", extra={
        'category': 'PROCESS'
    })
    return {"status": "cleared"}


@router.post("/remove-file", dependencies=[Depends(require_role("admin", "master"))])
async def remove_file(req: RemoveFileRequest):
    """Remove a single uploaded file by its stable file_id."""
    from ..main import uploaded_files, IMAGES_DIR
    result = _find_file_by_id(req.file_id)
    if result is None:
        raise HTTPException(status_code=404, detail="File not found")
    idx, entry = result
    # Delete from disk if it exists
    filepath = Path(entry.get("filepath", ""))
    if filepath.exists():
        try:
            filepath.unlink()
        except OSError:
            pass
        # Also clean up generated page images
        img_dir = IMAGES_DIR / filepath.stem
        if img_dir.exists():
            shutil.rmtree(str(img_dir), ignore_errors=True)
    uploaded_files.pop(idx)
    logger.info("File removed", extra={
        'category': 'PROCESS',
        'file_id': req.file_id
    })
    return {"status": "removed", "file_id": req.file_id}


@router.get("/next-file", dependencies=[Depends(require_role("admin", "master"))])
async def next_file(file_id: str | None = None, file_index: int = -1):
    """Get next file for processing, or pages for a specific file."""
    from ..main import uploaded_files, IMAGES_DIR
    
    # If file_id is provided, return pages for that file (used by review step)
    if file_id:
        result = _find_file_by_id(file_id)
        if result:
            idx, entry = result
            filepath = Path(entry.get("filepath", ""))
            img_dir = IMAGES_DIR / filepath.stem
            pages = []
            if img_dir.is_dir():
                pages = sorted([f"/images/{filepath.stem}/{p.name}" for p in img_dir.glob("page_*.png")])
            return {"file_id": file_id, "file_index": idx, "filename": entry["filename"], "pages": pages}
    
    # Legacy: if file_index is provided, return pages for that file
    if file_index >= 0 and file_index < len(uploaded_files):
        entry = uploaded_files[file_index]
        filepath = Path(entry.get("filepath", ""))
        img_dir = IMAGES_DIR / filepath.stem
        pages = []
        if img_dir.is_dir():
            pages = sorted([f"/images/{filepath.stem}/{p.name}" for p in img_dir.glob("page_*.png")])
        return {"file_id": entry.get("file_id"), "file_index": file_index, "filename": entry["filename"], "pages": pages}
    
    # Otherwise, find next uploaded file
    for i, entry in enumerate(uploaded_files):
        if entry["status"] == "uploaded":
            return {"file_id": entry.get("file_id"), "file_index": i, "filename": entry["filename"]}
    return {"file_id": None, "file_index": -1}


# ─── Process ──────────────────────────────────────────────

@router.post("/process-stream", dependencies=[Depends(require_role("admin", "master"))])
async def process_stream(req: ProcessRequest):
    """Process a file with streaming progress updates (SSE)."""
    from ..main import uploaded_files
    from fastapi.responses import StreamingResponse
    
    resolved = _resolve_file(file_id=req.file_id, file_index=req.file_index)
    if resolved is None:
        raise HTTPException(status_code=404, detail="File not found")
    
    idx, entry = resolved
    filepath = entry.get("filepath", "")
    if not filepath or not Path(filepath).exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    async def generate():
        """Yield SSE messages as processing progresses."""
        from ..parser import parse_file_with_ocr
        from ..extraction import extract_items_async
        from ..main import process_lock, IMAGES_DIR
        
        import time
        start_time = time.time()
        
        def send(msg):
            return f"data: {json.dumps(msg)}\n\n"
        
        # Try to acquire the processing lock (non-blocking)
        # Only one file can be processed at a time across all users
        if process_lock.locked():
            yield send({"type": "error", "message": "Another user is currently processing a file. Try again later."})
            return
        
        await process_lock.acquire()

        try:
            # Signal parsing started
            yield send({"type": "progress", "percent": 0, "page": 0, "total": 1, "message": "Parsing file..."})

            try:
                parse_result = await parse_file_with_ocr(filepath)
            except Exception as e:
                logger.error("Parse failed", extra={
                    'category': 'PROCESS',
                    'file': entry["filename"],
                    'file_id': entry.get("file_id"),
                    'error': str(e)
                })
                yield send({"type": "error", "message": f"Parse failed: {e}"})
                return

            if parse_result.get("error"):
                logger.error("Parse failed", extra={
                    'category': 'PROCESS',
                    'file': entry["filename"],
                    'file_id': entry.get("file_id"),
                    'error': parse_result["error"]
                })
                yield send({"type": "error", "message": parse_result["error"]})
                return

            # Pass pdf_path so extraction router can use Vision LLM for scanned PDFs
            parse_result["pdf_path"] = str(filepath)

            num_pages = parse_result.get("num_pages", 1)

            # Clean up stale page images from a previous (cancelled) run
            img_dir = IMAGES_DIR / Path(filepath).stem
            if img_dir.is_dir():
                shutil.rmtree(img_dir)

            # Generate page images for the review PDF viewer
            try:
                page_images = _generate_page_images(Path(filepath))
                entry["pages"] = page_images
            except Exception:
                logger.warning("Page image generation failed for %s", entry.get("filename"), exc_info=True)
                page_images = []

            # Send progress for each page (synthetic since parsing is already done)
            for page in range(1, num_pages + 1):
                percent = int((page / num_pages) * 80)
                yield send({"type": "progress", "percent": percent, "page": page, "total": num_pages, "message": f"Processing page {page}/{num_pages}..."})
                await asyncio.sleep(0.05)

            # Extraction
            yield send({"type": "progress", "percent": 80, "page": num_pages, "total": num_pages, "message": "Extracting items..."})

            try:
                from ..utils import get_config_data
                cfg = get_config_data()
                extraction_enabled = cfg.get("extraction_enabled", True)

                result = await extract_items_async(
                    parse_result,
                )
            except Exception as e:
                logger.error("Extraction failed", extra={
                    'category': 'PROCESS',
                    'file': entry["filename"],
                    'file_id': entry.get("file_id"),
                    'error': str(e)
                })
                yield send({"type": "error", "message": f"Extraction failed: {e}"})
                return

            # Calculate processing time
            processing_time = round(time.time() - start_time, 2)

            # Log successful processing
            logger.info("Processing complete", extra={
                'category': 'PROCESS',
                'file': entry["filename"],
                'file_id': entry.get("file_id"),
                'method': result.extraction_method,
                'items': len(result.items),
                'time': f"{processing_time}s",
                'warnings': len(result.warnings)
            })

            # Send page_done for each page with item counts
            items_per_page = {}
            for item in result.items:
                page = item.get("page", 1)
                items_per_page[page] = items_per_page.get(page, 0) + 1

            for page in range(1, num_pages + 1):
                items_found = items_per_page.get(page, 0)
                percent = int((page / num_pages) * 100)
                yield send({"type": "page_done", "percent": percent, "page": page, "total": num_pages, "items_found": items_found})

            # Send done message with full result
            yield send({
                "type": "done",
                "data": {
                    "filename": entry["filename"],
                    "items": result.items,
                    "supplier": result.supplier,
                    "date": result.date,
                    "currency": result.currency,
                    "document_type": result.document_type,
                    "extraction_method": result.extraction_method,
                    "warnings": result.warnings,
                }
            })
        finally:
            process_lock.release()
    
    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


# ─── Confirm / Save ───────────────────────────────────────

@router.post("/confirm", dependencies=[Depends(require_role("admin", "master"))])
async def confirm(req: ConfirmRequest):
    """Save processed data to database."""
    from ..main import uploaded_files, ARCHIVE_DIR, IMAGES_DIR
    
    resolved = _resolve_file(file_id=req.file_id, file_index=req.file_index)
    if resolved is None:
        raise HTTPException(status_code=404, detail="File not found")
    
    idx, entry = resolved
    
    data = req.data
    items = data.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="Cannot save: at least one item is required")
    supplier = data.get("supplier", "")
    quotation_date = data.get("date", "")
    currency = items[0].get("currency", "") if items else ""
    document_type = data.get("document_type", "unknown")
    extraction_method = data.get("extraction_method", "llm_first")
    
    # Insert into database
    with get_db() as db:
        db.execute(
            "INSERT INTO quotations (filename, supplier, quotation_date, currency, items, document_type, extraction_method) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (entry["filename"], supplier, quotation_date,
             currency, json.dumps(items), document_type, extraction_method)
        )
        last_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    
    # Move PDF to archive
    src = Path(entry["filepath"])
    if src.exists():
        dst = ARCHIVE_DIR / src.name
        shutil.move(str(src), str(dst))
    
    # Clean up images
    file_stem = Path(entry["filename"]).stem
    img_dir = IMAGES_DIR / file_stem
    if img_dir.exists():
        shutil.rmtree(str(img_dir))
    
    entry["status"] = "saved"
    
    logger.info("Quotation saved", extra={
        'category': 'PROCESS',
        'file': entry["filename"],
        'file_id': req.file_id,
        'db_id': last_id,
        'supplier': supplier,
        'document_type': document_type,
        'items': len(items)
    })
    
    return {"status": "saved", "id": last_id}


@router.post("/skip", dependencies=[Depends(require_role("admin", "master"))])
async def skip(req: ProcessRequest):
    """Skip current file."""
    from ..main import uploaded_files
    resolved = _resolve_file(file_id=req.file_id, file_index=req.file_index)
    if resolved is None:
        raise HTTPException(status_code=404, detail="File not found")
    idx, entry = resolved
    entry["status"] = "skipped"
    logger.info("File skipped", extra={
        'category': 'PROCESS',
        'file': entry["filename"],
        'file_id': req.file_id
    })
    return {"status": "skipped"}


# ─── Delete / Update ──────────────────────────────────────

@router.post("/delete", dependencies=[Depends(require_role("admin", "master"))])
async def delete(req: DeleteRequest):
    """Delete quotations by ID."""
    import shutil
    from pathlib import Path
    
    if not req.ids:
        return {"status": "nothing to delete"}
    
    # Get filenames before deleting
    placeholders = ",".join("?" * len(req.ids))
    with get_db(readonly=True) as db:
        rows = db.execute(
            f"SELECT id, filename FROM quotations WHERE id IN ({placeholders})", req.ids
        ).fetchall()
    
    if not rows:
        return {"status": "nothing to delete", "detail": "No matching quotations found"}
    
    # Delete from database
    with get_db() as db:
        cur = db.execute(f"DELETE FROM quotations WHERE id IN ({placeholders})", req.ids)
        entries_deleted = cur.rowcount
    
    # Delete archived PDFs and image directories
    from ..main import ARCHIVE_DIR, IMAGES_DIR
    files_deleted = 0
    for row in rows:
        filename = row["filename"]
        if not filename:
            continue
        
        # Delete archive PDF
        archive_path = ARCHIVE_DIR / filename
        try:
            if archive_path.exists():
                archive_path.unlink()
                files_deleted += 1
        except OSError:
            pass
        
        # Delete image directory
        img_dir = IMAGES_DIR / Path(filename).stem
        try:
            if img_dir.exists():
                shutil.rmtree(img_dir, ignore_errors=True)
        except OSError:
            pass
    
    logger.info("Quotations deleted", extra={
        'category': 'PROCESS',
        'ids': str(req.ids),
        'count': entries_deleted
    })
    
    return {"status": "deleted", "count": entries_deleted, "files_deleted": files_deleted}


@router.post("/update", dependencies=[Depends(require_role("admin", "master"))])
async def update(req: UpdateRequest):
    """Update a quotation."""
    data = req.data
    items = data.get("items", [])
    supplier = data.get("supplier", "")
    quotation_date = items[0].get("date", "") if items else ""
    currency = data.get("currency", "")
    document_type = data.get("document_type", "unknown")
    
    with get_db() as db:
        cur = db.execute(
            "UPDATE quotations SET supplier=?, quotation_date=?, currency=?, items=?, document_type=? WHERE id=?",
            (supplier, quotation_date, currency, json.dumps(items), document_type, req.id)
        )
    
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    logger.info("Quotation updated", extra={
        'category': 'PROCESS',
        'quotation_id': req.id,
        'supplier': supplier
    })
    
    return {"status": "updated"}


# ─── Export / Import ──────────────────────────────────────

@router.get("/export", dependencies=[Depends(require_role("admin", "master"))])
async def export_db():
    """Export all quotations as a zip file."""
    from ..main import ARCHIVE_DIR
    
    with get_db(readonly=True) as db:
        rows = db.execute("SELECT * FROM quotations ORDER BY created_at DESC").fetchall()
    
    data = []
    for r in rows:
        d = dict(r)
        if isinstance(d.get("items"), str):
            try:
                d["items"] = json.loads(d["items"])
            except (json.JSONDecodeError, TypeError):
                d["items"] = []
        data.append(d)
    
    # Create zip file
    zip_fd, zip_path = tempfile.mkstemp(suffix='.zip')
    os.close(zip_fd)
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            quotations_json = json.dumps({"quotations": data, "count": len(data)}, indent=2)
            zf.writestr("quotations.json", quotations_json)
            # Embed SHA256 checksum for integrity verification on import
            sha = hashlib.sha256(quotations_json.encode()).hexdigest()
            zf.writestr("quotations.json.sha256", sha)
            if ARCHIVE_DIR.exists():
                for pdf_file in ARCHIVE_DIR.glob("*.pdf"):
                    zf.write(pdf_file, f"archive/{pdf_file.name}")
        
        zip_size = os.path.getsize(zip_path)
        filename = f"quodb_backup_{__import__('datetime').datetime.now().strftime('%Y-%m-%d')}.zip"
        
        logger.info("Database exported", extra={
            'category': 'ADMIN',
            'row_count': len(data),
            'zip_size': f"{zip_size / 1024:.1f}KB"
        })
        
        def stream_file():
            with open(zip_path, "rb") as f:
                while chunk := f.read(65536):
                    yield chunk
            os.unlink(zip_path)
        
        return StreamingResponse(
            stream_file(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(zip_size)
            }
        )
    except Exception:
        logger.warning("Zip download stream failed", exc_info=True)
        if os.path.exists(zip_path):
            os.unlink(zip_path)
        raise


@router.post("/import/upload", dependencies=[Depends(require_role("admin", "master"))])
async def import_upload(file: UploadFile = File(...)):
    """Import quotations from a JSON or ZIP file."""
    content = await file.read()
    quotations = []
    pdf_restored = 0
    pdf_restored_paths = []  # Track restored PDFs for orphan cleanup on failure
    integrity_warning = None
    
    if file.filename.endswith(".zip"):
        import io
        try:
            with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                if "quotations.json" in zf.namelist():
                    # Verify SHA256 checksum if present (backward compatible)
                    if "quotations.json.sha256" in zf.namelist():
                        expected_sha = zf.read("quotations.json.sha256").decode().strip()
                        actual_sha = hashlib.sha256(zf.read("quotations.json")).hexdigest()
                        if expected_sha != actual_sha:
                            logger.warning("Import rejected: checksum mismatch", extra={
                                'category': 'ADMIN', 'file': file.filename,
                            })
                            return JSONResponse(status_code=400, content={
                                "error": "File integrity check failed — the data has been modified since export"
                            })
                    else:
                        logger.warning("Import ZIP has no SHA checksum — integrity not verified", extra={
                            'category': 'ADMIN', 'file': file.filename,
                        })
                        integrity_warning = "No integrity checksum found in ZIP — file not verified"
                    data = json.loads(zf.read("quotations.json"))
                    quotations = data.get("quotations", [])
                from ..main import ARCHIVE_DIR
                ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
                for name in zf.namelist():
                    if name.startswith("archive/") and name.endswith(".pdf"):
                        pdf_name = Path(name.split("/", 1)[1]).name
                        pdf_data = zf.read(name)
                        pdf_path = ARCHIVE_DIR / pdf_name
                        with open(pdf_path, "wb") as f:
                            f.write(pdf_data)
                        pdf_restored += 1
                        pdf_restored_paths.append(pdf_path)
        except zipfile.BadZipFile:
            return JSONResponse(status_code=400, content={"error": "Invalid zip file"})
    elif file.filename.endswith(".json"):
        try:
            data = json.loads(content)
            quotations = data.get("quotations", [])
        except json.JSONDecodeError:
            return JSONResponse(status_code=400, content={"error": "Invalid JSON"})
    else:
        return JSONResponse(status_code=400, content={"error": "Use .zip or .json"})
    
    if not quotations:
        # Clean up any PDFs restored before the failure (orphan prevention)
        for p in pdf_restored_paths:
            if p.exists():
                p.unlink()
        return JSONResponse(status_code=400, content={"error": "No quotations found"})
    
    imported = 0
    skipped = 0
    with get_db() as db:
        for q in quotations:
            supplier = q.get("supplier", "")
            quotation_date = q.get("quotation_date", "")
            document_type = q.get("document_type", "unknown")
            items = q.get("items", [])
            if isinstance(items, str):
                try:
                    items = json.loads(items)
                except (json.JSONDecodeError, TypeError):
                    items = []
            if not items:
                skipped += 1
                continue
            if not quotation_date:
                quotation_date = items[0].get("date", "")
            db.execute(
                "INSERT INTO quotations (filename, supplier, quotation_date, items, document_type) VALUES (?, ?, ?, ?, ?)",
                (q.get("filename", "imported.pdf"), supplier, quotation_date,
                 json.dumps(items), document_type)
            )
            imported += 1
    
    if imported == 0 and skipped > 0:
        # Clean up any PDFs restored before the failure (orphan prevention)
        for p in pdf_restored_paths:
            if p.exists():
                p.unlink()
        return JSONResponse(status_code=400, content={
            "error": f"Import failed: {skipped} quotation(s) found but all had no items"
        })
    
    logger.info("Database imported", extra={
        'category': 'ADMIN',
        'file': file.filename,
        'imported': imported,
        'skipped': skipped,
        'pdfs_restored': pdf_restored
    })
    
    result = {"status": "imported", "count": imported, "pdfs_restored": pdf_restored}
    if skipped > 0:
        result["skipped"] = skipped
    if integrity_warning:
        result["warning"] = integrity_warning
    return result
